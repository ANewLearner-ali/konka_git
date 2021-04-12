"""
平台可靠性工具excel报告处理通用模块
"""
import logging
import os

import openpyxl
from openpyxl.styles.alignment import Alignment
from openpyxl.styles.borders import Side, Border
from openpyxl.utils import cell as cell_utils


def create_report(excel: str):
    """
    创建报告
    :param excel:
    :return:
    """
    wb = openpyxl.Workbook()
    wb.worksheets[0].title = '中心主题'
    sheet = wb.worksheets[0]
    sheet['A1'] = '可靠性结果输出'
    sheet['A2'] = '压测项目'
    sheet['B2'] = '配置环境'
    sheet['C2'] = '总计时/总计次'
    sheet['D2'] = '检测结果'
    sheet['F2'] = '失败'
    sheet.merge_cells('A1:F1')
    sheet.merge_cells('D2:E2')
    wb.save(excel)


def get_start_row(excel: str):
    return openpyxl.load_workbook(excel).worksheets[0].max_row + 1


def init_main_sheet(excel, first, second):
    """
    初始化主sheet
    :param excel:
    :param first:
    :param second:
    :return:
    """
    wb = openpyxl.load_workbook(excel)
    sheet = wb.worksheets[0]
    row = sheet.max_row + 1
    sheet['A' + str(row)] = first
    sheet['B' + str(row)] = second
    wb.save(excel)


def upgrade_summary(excel: str, start_row: int, count: str, summary: tuple, record_times=False, count_text=''):
    """
    更新检测统计结果
    """
    wb = openpyxl.load_workbook(excel)
    sheet = wb.worksheets[0]
    sheet['C' + str(start_row)] = count
    summary, result = summary
    for i, key in enumerate(summary):
        sheet['D' + str(i + start_row)] = str(summary[key][0])
        if isinstance(summary[key][1], (int, float)):
            sheet['E' + str(i + start_row)] = str(round(summary[key][1] * 100, 2)) + '%'
            if record_times:
                if not result[key][1]:
                    ori = sheet['F' + str(i + start_row)].value if sheet['F' + str(i + start_row)].value else ''
                    sheet['F' + str(i + start_row)] = ori + count_text + '、'
        else:
            sheet['E' + str(i + start_row)] = str(summary[key][1])
    wb.save(excel)


def merge_summary(excel: str, start_row: int):
    """
    按照报告输出要求合并单元格
    :param excel:
    :param start_row:
    :return:
    """
    if not os.path.isfile(excel):
        return
    wb = openpyxl.load_workbook(excel)
    sheet = wb.worksheets[0]
    max_row = sheet.max_row
    sheet.merge_cells('A' + str(start_row) + ':' + 'A' + str(max_row))
    sheet.merge_cells('B' + str(start_row) + ':' + 'B' + str(max_row))
    sheet.merge_cells('C' + str(start_row) + ':' + 'C' + str(max_row))
    wb.save(excel)


def optimize(excel):
    """
    测试报告优化，包括文字居中和边框设置
    :param excel:
    :return:
    """
    if not os.path.isfile(excel):
        logging.warning(f'optimize excel fail , file : {excel}')
        return
    wwb = openpyxl.load_workbook(excel)
    side = Side(style='thin', color='FF000000')
    for sheet in wwb.worksheets:
        for row in sheet.iter_rows():
            for cell in row:
                cell.border = Border(
                    left=side, right=side, top=side, bottom=side)
                cell.alignment = Alignment(horizontal='left', vertical='center', wrapText=True)
    wwb.save(excel)


#######################################
# Performance
#######################################
def _p_tittle(sheet, processes: list):
    """
    性能报告tittle设置
    :param sheet:
    :param processes:
    :return:
    """
    cur_letter = 'C'

    for process in processes:
        if cur_letter == 'C':
            sheet[cur_letter + '1'] = process
            sheet[cur_letter + '2'] = '内存'
            next_letter = cell_utils.get_column_letter(cell_utils.column_index_from_string(cur_letter) + 1)
            sheet[next_letter + '2'] = 'CPU'
            sheet.merge_cells(cur_letter + '1:' + next_letter + '1')
            cur_letter = cell_utils.get_column_letter(cell_utils.column_index_from_string(next_letter) + 1)
        else:
            sheet[cur_letter + '1'] = process
            sheet[cur_letter + '2'] = 'PID'
            mid_letter = cell_utils.get_column_letter(cell_utils.column_index_from_string(cur_letter) + 1)
            sheet[mid_letter + '2'] = '内存'
            end_letter = cell_utils.get_column_letter(cell_utils.column_index_from_string(mid_letter) + 1)
            sheet[end_letter + '2'] = 'CPU'
            sheet.merge_cells(cur_letter + '1:' + end_letter + '1')
            cur_letter = cell_utils.get_column_letter(cell_utils.column_index_from_string(end_letter) + 1)


def p_create_report(excel: str, processes: list):
    """
    性能测试报告创建
    :param excel:
    :param processes:
    :return:
    """
    wb = openpyxl.Workbook()
    wb.worksheets[0].title = '性能数据'
    sheet = wb.worksheets[0]
    sheet['A1'] = '时间'
    sheet['B1'] = '模式'
    sheet.merge_cells('A1:A2')
    sheet.merge_cells('B1:B2')
    _p_tittle(sheet, processes)
    wb.save(excel)


def _p_add_pf(sheet, pf: list, row):
    cur_letter = 'C'
    for p in pf:
        if cur_letter == 'C':
            sheet[cur_letter + row] = p[1]
            next_letter = cell_utils.get_column_letter(cell_utils.column_index_from_string(cur_letter) + 1)
            sheet[next_letter + row] = p[2]
            cur_letter = cell_utils.get_column_letter(cell_utils.column_index_from_string(next_letter) + 1)
        else:
            sheet[cur_letter + row] = p[0]
            mid_letter = cell_utils.get_column_letter(cell_utils.column_index_from_string(cur_letter) + 1)
            sheet[mid_letter + row] = p[1]
            end_letter = cell_utils.get_column_letter(cell_utils.column_index_from_string(mid_letter) + 1)
            sheet[end_letter + row] = p[2]
            cur_letter = cell_utils.get_column_letter(cell_utils.column_index_from_string(end_letter) + 1)


def p_append(excel: str, info: dict):
    """
    追加一行性能数据
    :param excel:
    :param info:
    :return:
    """
    wb = openpyxl.load_workbook(excel)
    sheet = wb.worksheets[0]
    row = str(sheet.max_row + 1)
    sheet['A' + row] = info['tv_time']
    sheet['B' + row] = info['mode']
    _p_add_pf(sheet, info['pf'], row)
    wb.save(excel)
